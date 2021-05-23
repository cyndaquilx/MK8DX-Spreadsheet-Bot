import discord
from discord.ext import commands

import aiosqlite
import gspread_asyncio
from oauth2client.service_account import ServiceAccountCredentials

import openpyxl
import excel2img
import json
from datetime import date

#import asyncio

from constants import (channels, ranks, getRank,
                       SH_KEY, updateCols, getCols,
                       peakColumn, rowOffset, colOffset,
                       sheet_start_rows, table_start_rows,
                       rowcol_to_a1, get_strike_info,
                       place_MMRs, pen_row, pen_cols)

def get_creds():
    return ServiceAccountCredentials.from_json_keyfile_name(
        "credentials.json",
        [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive",
            "https://www.googleapis.com/auth/spreadsheets",
        ],
    )
agcm = gspread_asyncio.AsyncioGspreadClientManager(get_creds)

def findmember(ctx, name, roleid):
    #print(name)
    members = ctx.guild.members
    role = ctx.guild.get_role(roleid)
    def pred(m):
        #print(m.name)
        if m.nick is not None:
            if m.nick.lower() == name.lower():
                return True
        if m.name.lower() != name.lower():
            return False
        if role not in m.roles:
            return False
        return True
        #return m.nick.lower() == name.lower() or m.name.lower() == name.lower()
    return discord.utils.find(pred, members)

class Updating(commands.Cog):
    def __init__(self, bot):
        self.bot = bot
        with open('./config.json', 'r') as cjson:
            self.config = json.load(cjson)

    

    @commands.max_concurrency(number=1, wait=True)
    @commands.group(aliases=['u'])
    @commands.has_any_role("Administrator", "Updater", "Staff-S")
    async def update(self, ctx):
        if ctx.invoked_subcommand is None:
            return

    @commands.max_concurrency(number=1, wait=True)
    @update.command(aliases=['a'])
    @commands.has_any_role("Administrator", "Updater", "Staff-S")
    async def approve(self, ctx, tableid:int, *, extraArgs=""):
        if ctx.guild.id != self.config["server"]:
            await ctx.send("You cannot use this command in this server!")
            return
        try:
            db = await aiosqlite.connect('updating.db')
            c = await db.cursor()
            await c.execute("SELECT * from tables WHERE tableid = ?", (tableid,))
            table = await c.fetchone()
            size = int(table[1])
            tier = table[2].upper()
            names = table[3].split(",")
            placements = table[4].split(",")
            messageid = table[6]
            #await ctx.send("%s\n%s\n%s\n%s"
            #               % (size, tier, names, placements))
        except:
            await ctx.send("Table couldn't be found")
            return
        finally:
            await db.close()
            
        arguments = extraArgs.split(";")
        instructions = {}
        if tier.upper() == "SQ":
            for i in range(12):
                instructions[i+1] = 0.75
                
        #processing optional arguments to give specific players multipliers
        elif len(arguments) > 0:
            try:
                instructions = await Updating.processInstructions(self, ctx, tier, arguments[0])
            except:
                return
        mults = []
        for i in range(12):
            if(i+1) in instructions.keys():
                mults.append(instructions[i+1])
            else:
                mults.append(1)

        races = 12
        if len(arguments) > 1:
            try:
                raceArg = int(arguments[1])
            except:
                await(await ctx.send("The number of races you entered is not an integer between 1-12; try again")).delete(delay=10)
                return
            if raceArg < 1 or raceArg > 12:
                await(await ctx.send("The number of races you entered is not an integer between 1-12; try again")).delete(delay=10)
                return
            races = raceArg
        #await ctx.send("%s\n%s" % (mults, races))
        try:
            await Updating.updateTable(self, ctx, size, tier, names, placements, mults, races, messageid, tableid)
        except Exception as e:
            #print(e)
            return
        try:
            db = await aiosqlite.connect('updating.db')
            c = await db.cursor()
            await c.execute("DELETE from tables WHERE tableid = ?", (tableid,))
            #table = await c.fetchone()
            #size = int(table[1])
            #tier = table[2].upper()
            #names = table[3].split(",")
            #placements = table[4].split(",")
            await db.commit()
            #await ctx.send("Successfully removed table %d from database" % tableid)
        except:
            await ctx.send("Database error removing table from approval queue")
            return
        finally:
            await db.close()
        channel = ctx.guild.get_channel(channels[tier.upper()])
        try:
            reactMsg = await channel.fetch_message(messageid)
            CHECK_BOX = "\U00002611"
            await reactMsg.add_reaction(CHECK_BOX)
        except Exception as e:
            #print(e)
            return

    @commands.max_concurrency(number=1, wait=True)
    @update.command(aliases=['d'])
    @commands.has_any_role("Administrator", "Updater", "Staff-S")
    async def deny(self, ctx, tableid:int):
        if ctx.guild.id != self.config["server"]:
            await ctx.send("You cannot use this command in this server!")
            return
        try:
            db = await aiosqlite.connect('updating.db')
            c = await db.cursor()
            await c.execute("DELETE from tables WHERE tableid = ?", (tableid,))
            await db.commit()
            await ctx.send("Removed table %d from approval queue" % tableid)
        except:
            await ctx.send("Database error removing table from approval queue")
            return
        finally:
            await db.close()
        
        
        
    @commands.max_concurrency(number=1, wait=True)
    @update.command(aliases=['t'])
    @commands.has_any_role("Administrator", "Updater", "Staff-S")
    async def text(self, ctx, size:int, tier, *, args):
        if ctx.guild.id != self.config["server"]:
            await ctx.send("You cannot use this command in this server!")
            return
        
        arguments = args.split(";")
        if len(arguments) < 2:
            await(await ctx.send("Please enter names and placements, separated by ;")).delete(delay=15)
            return
        if size not in [1, 2, 3, 4, 6]:
            await(await ctx.send("Please enter a valid format: 1, 2, 3, 4, or 6")).delete(delay=10)
            return
        if tier.upper() not in channels.keys():
            await(await ctx.send("Please enter a valid tier: X, S, A, B, C, D, or E")).delete(delay=10)
            return
        
        players = arguments[0].split(",")
        placements = arguments[1].strip().split(" ")
        names = [player.strip() for player in players]
        if len(names) != 12:
            await(await ctx.send("Please type exactly 12 player names")).delete(delay=10)
            return
        if len(placements) != (12/size):
            await(await ctx.send("Please type exactly %d placements after the comma"
                           % (12/size))).delete(delay=10)
            return
        if len(set(names)) < len(names):
            await(await ctx.send("There are duplicates in your input; try again")).delete(delay=10)
            return
    
        for placement in placements:
            try:
                intp = int(placement)
            except:
                await(await ctx.send("Your argument %s is not a placement between 1 and %d!"
                               % (placement, int(12/size)))).delete(delay=10)
                return
            if int(placement) not in range(1, int(12/size)+1):
                await(await ctx.send("Your argument %s is not a placement between 1 and %d!"
                               % (placement, int(12/size)))).delete(delay=10)
                return

        #processing optional arguments to give specific players multipliers
        instructions = {}
        if tier.upper() == "SQ":
            for i in range(12):
                instructions[i+1] = 0.75
        elif len(arguments) > 2:
            try:
                instructions = await Updating.processInstructions(self, ctx, tier, arguments[2])
            except:
                return
            
        mults = []
        for i in range(12):
            if(i+1) in instructions.keys():
                mults.append(instructions[i+1])
            else:
                mults.append(1)
        
        #processing optional race count argument
        races = 12
        if len(arguments) > 3:
            try:
                raceArg = int(arguments[3])
            except:
                await(await ctx.send("The number of races you entered is not an integer between 1-12; try again")).delete(delay=10)
                return
            if raceArg < 1 or raceArg > 12:
                await(await ctx.send("The number of races you entered is not an integer between 1-12; try again")).delete(delay=10)
                return
            races = raceArg
        try:
            await Updating.updateTable(self, ctx, size, tier, names, placements, mults, races)
        except Exception as e:
            #print(e)
            return
        #await ctx.send("Done")
            

    async def updateTable(self, ctx, size:int, tier, names, placements, mults, races, messageid=0, tableid=0):
        msg = await ctx.send("Working...")
        agc = await agcm.authorize()
        sh = await agc.open_by_key(SH_KEY)
        botSheet = await sh.worksheet("Bot")
        pHistory = await sh.worksheet("Player History")
        idSheet = await sh.worksheet("ID")
        
        channel = ctx.guild.get_channel(channels[tier.upper()])
        start = sheet_start_rows[size]
              

        #first we have to add all the table data to the spreadsheet
        #in order to get MMR changes, as well as check that all the
        #player names are correct
        updateCells = [{
            'range': "%s%d:%s%d" % (updateCols[0], start,
                                    updateCols[0], start+11),
            'values': [[name] for name in names]}, {
            'range': "%s%d:%s%d" % (updateCols[1], start,
                                    updateCols[1], start+int(12/size)-1),
            'values': [[int(placement)] for placement in placements]}, {
            'range': "%s%d:%s%d" % (updateCols[2], start,
                                    updateCols[2], start+11),
            'values': [[mult] for mult in mults]}
            ]
        updateCells.append({'range': "C%d" % (start+12),
                            'values': [[races]]})
        await botSheet.batch_update(updateCells)

        gotBatch = await botSheet.batch_get(["%s%d:%s%d" % (getCols[0], start,
                                                            getCols[1], start+11)])
        peakMMRs = []
        oldMMRs = []
        mmrChanges = []
        newMMRs = []
        rowNums = []
        colNums = []
        goodNames = []
        for i in range(12):
            peakMMRs.append(gotBatch[0][i][0])
            oldMMRs.append(gotBatch[0][i][1])
            mmrChanges.append(int(gotBatch[0][i][2]))
            newMMRs.append(int(gotBatch[0][i][3]))
            rowNums.append(gotBatch[0][i][4])
            colNums.append(int(gotBatch[0][i][5]))
            goodNames.append(gotBatch[0][i][6])

        idCell = await idSheet.acell('A1')
        idNum = int(idCell.value)

        updateCells = []
        peakChanges = []
        errors = ""
        for i in range(12):
            if rowNums[i] == "#N/A" or oldMMRs[i] == "N/A":
                #await(await ctx.send("Player %s is not on the sheet; check your input" % names[i])).delete(delay=10)
                #await msg.delete()
                #raise Exception()
                errors += "Player %s is not on the sheet; check your input\n" % names[i]
            if colNums[i] >= 399:
                #await(await ctx.send("Player %s needs to be archived, which is not supported by this bot; please update this table with the sheet script"
                #               % names[i])).delete(delay=20)
                #await msg.delete()
                #raise Exception()
                errors += ("Player %s needs to be archived, which is not supported by this bot; please update this table with the sheet script\n"
                           % names[i])
            if oldMMRs[i] == "Placement":
                #await(await ctx.send("Player %s needs to be given Placement MMR; please give them a base MMR on the sheet"
                #               % names[i])).delete(delay=15)
                #await msg.delete()
                #raise Exception()
                errors += ("Player %s needs to be given Placement MMR; please give them a base MMR on the sheet\n"
                           % names[i])
        if len(errors) > 0:
            await ctx.send(errors)
            await msg.delete()
            raise Exception()
        for i in range(12):
            if peakMMRs[i] == "N/A":
                if colNums[i] >= 4:
                    peakCell = {'range': "%s%d" % (peakColumn,
                                                   int(rowNums[i])+rowOffset),
                                'values': [[newMMRs[i]]]}
                    updateCells.append(peakCell)
                    peakChanges.append([str(int(rowNums[i])+rowOffset), "N/A"])
            elif newMMRs[i] > int(peakMMRs[i]):
                peakCell = {'range': "%s%d" % (peakColumn,
                                               int(rowNums[i])+rowOffset),
                            'values': [[newMMRs[i]]]}
                updateCells.append(peakCell)
                peakChanges.append([str(int(rowNums[i])+rowOffset), str(peakMMRs[i])])
            while int(oldMMRs[i]) + mmrChanges[i] < 0:
                mmrChanges[i] += 1
            cellA1 = rowcol_to_a1(int(rowNums[i])+rowOffset, colNums[i]+colOffset)
            cell = {'range': cellA1,
                    'values': [[mmrChanges[i]]]}
            updateCells.append(cell)
        idNum += 1
        await idSheet.update_cell(1, 1, idNum)

        wb = openpyxl.load_workbook("Updating.xlsx")
        ws = wb["Sheet1"]

        start = table_start_rows[size]

        index = 0
        tierCell = ws["E%s" % (start-2)]
        if tier.upper() != "SQ":
            tierCell.value = "Tier %s" % tier.upper()
        else:
            tierCell.value = "Squad Queue"
        for i in range(int(12/size)):
            placeCell = ws["C%s" % (start+index)]
            placeCell.value = int(placements[i])
            for j in range(size):
                ij = i*size + j
                if peakMMRs[ij] != "N/A":
                    peakMMRs[ij] = int(peakMMRs[ij])
                    
                peakCell = ws["B%s" % (start+index)]
                peakCell.value = peakMMRs[ij]
                
                playerCell = ws["D%s" % (start+index)]
                playerCell.value = goodNames[ij]

                oldMMRcell = ws["E%s" % (start+index)]
                oldMMRcell.value = int(oldMMRs[ij])

                changeCell = ws["F%s" % (start+index)]
                changeCell.value = int(mmrChanges[ij])

                newMMRcell = ws["G%s" % (start+index)]
                newMMRcell.value = int(newMMRs[ij])
                index += 1
            if size > 1 and i+1 < 12/size:
                index += 1
        racesCell = ws["D%s" % (start+index)]
        racesCell.value = races
        idCell = ws["H%s" % (start+index)]
        idCell.value = idNum

        imgrange = "Sheet1!C%d:H%d" % (start-2, start+index)

        wb.save("Updating.xlsx")
        excel2img.export_img("Updating.xlsx", "test.png", "", imgrange)

        #processing any potential rank changes
        #print("checking rank changes...")
        rankchanges = ""
        for i in range(12):
            rank1 = getRank(int(oldMMRs[i]))
            rank2 = getRank(int(newMMRs[i]))
            if rank1 != rank2:
                #member = ctx.guild.get_member_named(goodNames[i])
                member = findmember(ctx, goodNames[i], ranks[rank1]["roleid"])
                if member is not None:
                    memName = member.mention
                else:
                    memName = goodNames[i]
                rankchanges += ("%s -> %s\n"
                                % (memName, ranks[rank2]["emoji"]))
                role1 = ctx.guild.get_role(ranks[rank1]["roleid"])
                role2 = ctx.guild.get_role(ranks[rank2]["roleid"])
                if member is not None:
                    if role1 in member.roles:
                        await member.remove_roles(role1)
                    if role2 not in member.roles:
                        await member.add_roles(role2)
        #print("done checking")
        await pHistory.batch_update(updateCells)
        await msg.delete()
        msg = "Table updated successfully"
        if ctx.channel != channel:
            msg += "; check %s to view" % channel.mention
        myMsg = await ctx.send(msg)

        if ctx.channel == channel:
            await myMsg.delete(delay=5)
            try:
                await ctx.message.delete()
            except:
                pass

        f = discord.File("test.png", filename="MMRTable.png")
        e = discord.Embed(title="MMR Table")
        e.add_field(name="ID", value=str(idNum))
        e.add_field(name="Tier", value=tier.upper())
        if messageid != 0 and tableid != 0:
            try:
                foundmsg = await channel.fetch_message(messageid)
            except:
                foundmsg = None
            if foundmsg is not None:
                submissionContent = "[%d](%s)" % (tableid, foundmsg.jump_url)
            else:
                submissionContent = "%d" % tableid
            e.add_field(name="Submission ID", value=submissionContent)
            
        e.add_field(name="Updated by", value=ctx.author.mention)
        lossMultStr = ""
        if tier.upper() != "SQ":
            for i in range(12):
                if mults[i] != 1:
                    lossMultStr += ("%.2fx MMR multiplier for %s\n"
                                   % (mults[i], goodNames[i]))
        if lossMultStr:
            e.add_field(name="Notes", value=lossMultStr)
        e.set_image(url="attachment://MMRTable.png")
        sentmsg = await channel.send(content=rankchanges,
                                     file=f, embed=e)
        rowNumStr = ",".join([str(rowNum) for rowNum in rowNums])
        colNumStr = ",".join([str(colNum) for colNum in colNums])
        peakChangesStr = ",".join([",".join(change) for change in peakChanges])
        msgid = sentmsg.id
        db_entry = (idNum, rowNumStr, colNumStr, peakChangesStr, msgid, tier.upper())
        #print(db_entry)
        try:
            db = await aiosqlite.connect('updating.db')
            c = await db.cursor()
            await c.execute("""INSERT INTO updated
                            (tableid, rowids, colids, peakChanges, msgid, tier)
                            VALUES (?,?,?,?,?,?)
                            """, db_entry)
            await db.commit()
        except Exception as e:
            #print(e)
            return
        finally:
            await db.close()

    @update.command(aliases=['u'])
    @commands.max_concurrency(number=1,wait=True)
    @commands.has_any_role("Administrator", "Updater", "Staff-S")
    async def undo(self, ctx, idNum: int):
        if ctx.guild.id != self.config["server"]:
            await ctx.send("You cannot use this command in this server!")
            return
        agc = await agcm.authorize()
        sh = await agc.open_by_key(SH_KEY)
        pHistory = await sh.worksheet("Player History")
        try:
            db = await aiosqlite.connect('updating.db')
            c = await db.cursor()
            await c.execute("SELECT * from updated WHERE tableid = ?", (idNum,))
            table = await c.fetchone()
            #print(table)
            rowids = table[1].split(",")
            colids = table[2].split(",")
            peakchanges = table[3].split(",")
            msgid = table[4]
            tier = table[5]
            clearedCells = []
            #print(table)
            for i in range(12):
                #print(rowcol_to_a1(int(rowids[i])+rowOffset, int(colids[i])+colOffset))
                clearCell = {'range': rowcol_to_a1(int(rowids[i])+rowOffset, int(colids[i])+colOffset),
                             'values': [['']]}
                clearedCells.append(clearCell)
            for i in range(int(len(peakchanges)/2)):
                if peakchanges[2*i+1] != "N/A":
                    oldpeak = int(peakchanges[2*i+1])
                else:
                    oldpeak = peakchanges[2*i+1]
                peakCell = {'range': "%s%d" % (peakColumn, int(peakchanges[2*i])),
                            'values': [[oldpeak]]}
                clearedCells.append(peakCell)
            await pHistory.batch_update(clearedCells)
            channel = ctx.guild.get_channel(channels[tier.upper()])
            await c.execute("DELETE from updated WHERE tableid = ?", (idNum,))
            await db.commit()
            try:
                msg = await channel.fetch_message(msgid)
                await msg.delete()
            except:
                pass
        except Exception as e:
            #print(e)
            await ctx.send("Table not found")
            return
        finally:
            await db.close()
        if channel != ctx.channel:
            await ctx.send("Done")
            
            

    @update.command(aliases=['s'])
    @commands.max_concurrency(number=1,wait=True)
    @commands.has_any_role("Administrator", "Updater", "Staff-S")
    async def strike(self, ctx, amount:int, tier, *, args):
        if ctx.guild.id != self.config["server"]:
            await ctx.send("You cannot use this command in this server!")
            return
        if amount < 0:
            await(await ctx.send("Please enter a positive amount")).delete(delay=10)
            return
        if tier.upper() not in channels.keys():
            await(await ctx.send("Please enter a valid tier")).delete(delay=10)
            return
        agc = await agcm.authorize()
        sh = await agc.open_by_key(SH_KEY)
        botSheet = await sh.worksheet("Bot")
        pHistory = await sh.worksheet("Player History")
        strikeSheet = await sh.worksheet("Strikes")
        channel = ctx.guild.get_channel(channels[tier.upper()])
        arguments = args.split(";")
        playername = arguments[0].strip()
        reason = ""
        if len(arguments) > 1:
            reason = arguments[1].strip()
        
        #updateCells = [{'range': pen_cells[0],
        #                'values': [[playername]]},
        #               {'range': pen_cells[1],
        #                'values': [[amount]]}]
        updateCells = [{'range': rowcol_to_a1(pen_row, pen_cols[0]),
                        'values': [[playername]]},
                       {'range': rowcol_to_a1(pen_row, pen_cols[1]),
                        'values': [[amount]]}]
        await botSheet.batch_update(updateCells)
        info = await botSheet.batch_get(["%s:%s" % (get_strike_info[0], get_strike_info[1])])

        strikes = info[0][0][4:7]
        pHrow = info[0][0][2]
        if pHrow == "#N/A":
            await(await ctx.send("This player is not on the sheet; check your input and try again")).delete(delay=10)
            return
        strikeRow = info[0][0][1]
        newRow = info[0][0][0]
        goodName = info[0][0][8]
        
        offset = 0
        for strike in strikes:
            if strike == "":
                break
            offset += 1
        if offset == 3:
            await ctx.send("This player already has 3 strikes; if you believe this isn't correct, please remove the extra strikes from the sheet")
            return
        updateCells = []
        #if the player isnt on the strike sheet, use the NewRow column
        #instead of StrikeRow
        if strikeRow == "Not found":
            rowUpdate = int(newRow)
            updateCells.append({'range': "A%d" % (rowUpdate),
                        'values': [[goodName]]})
        else:
            rowUpdate = int(strikeRow)
            
        pens = int(info[0][0][3])
        mmr = int(info[0][0][7])
        while mmr - amount < 0:
            amount -= 1
        pens += amount
        
        today = date.today()
        todaysdate = today.strftime("%m/%d/%y")
        m, d, y = map(int, todaysdate.split("/"))
        m += 1
        if m > 12:
            m-=12
            y+=1
        expireDate = "%d/%d/%d" % (m, d, y)

        updateCells.append({'range': rowcol_to_a1(rowUpdate, 4+offset),
                        'values': [[todaysdate]]})
        await strikeSheet.batch_update(updateCells)
        await pHistory.update_cell(pHrow, 4, pens)
        
        content = "Successfully added -%d and strike to %s" % (amount, goodName)
            
        e = discord.Embed(title="Strike + penalty added")
        e.add_field(name="Player", value=goodName, inline=False)
        e.add_field(name="Penalty", value="-%d MMR:\n%d -> %d" % (amount, mmr, mmr-amount), inline=False)
        
        strikecount = 0
        strikeData = ""
        for i in range(3):
            if info[0][0][4+i] == "":
                continue
            m, d, y = map(int, str(info[0][0][4+i]).split("/"))
            m+=1
            if m > 12:
                m-=12
                y+=1
            strikeData += ("\nStrike %d: expires on %d/%d/%d"
                           % (strikecount+1, m, d, y))
            strikecount += 1
        strikeData += ("\nStrike %d: expires on %s" % (strikecount+1, expireDate))
        strikeData = "%d/3 strikes\n%s" % (strikecount+1, strikeData)
        e.add_field(name="Strike info", value=strikeData, inline=False)

        if strikecount + 1 == 3:
            content += ("\n%s player %s has reached the strike limit and should be muted"
                        % (ctx.author.mention, goodName))
        msg = await ctx.send(content)
        if ctx.channel == channel:
            try:
                await ctx.message.delete()
            except:
                pass
            await msg.delete(delay=15)

        if reason != "":
            e.add_field(name="Reason", value=reason, inline=False)
        
        await channel.send(embed=e)

    @update.command(aliases=['pen'])
    @commands.max_concurrency(number=1,wait=True)
    @commands.has_any_role("Administrator", "Updater", "Staff-S")
    async def penalty(self, ctx, amount:int, tier, *, args):
        if ctx.guild.id != self.config["server"]:
            await ctx.send("You cannot use this command in this server!")
            return
        if amount < 0:
            await(await ctx.send("Please enter a positive amount")).delete(delay=10)
            return
        if tier.upper() not in channels.keys():
            await(await ctx.send("Please enter a valid tier")).delete(delay=10)
            return
        agc = await agcm.authorize()
        sh = await agc.open_by_key(SH_KEY)
        botSheet = await sh.worksheet("Bot")
        pHistory = await sh.worksheet("Player History")
        channel = ctx.guild.get_channel(channels[tier.upper()])
        arguments = args.split(";")
        playername = arguments[0].strip()
        reason = ""
        if len(arguments) > 1:
            reason = arguments[1].strip()
        #updateCells = [{'range': pen_cells[0],
        #                'values': [[playername]]},
        #               {'range': pen_cells[1],
        #                'values': [[amount]]}]
        updateCells = [{'range': rowcol_to_a1(pen_row, pen_cols[0]),
                        'values': [[playername]]},
                       {'range': rowcol_to_a1(pen_row, pen_cols[1]),
                        'values': [[amount]]}]
        await botSheet.batch_update(updateCells)
        info = await botSheet.batch_get(["%s:%s" % (get_strike_info[0], get_strike_info[1])])
        pHrow = info[0][0][2]
        if pHrow == "#N/A":
            await(await ctx.send("This player is not on the sheet; check your input and try again")).delete(delay=15)
            return
        pens = int(info[0][0][3])
        mmr = info[0][0][7]
        goodName = info[0][0][8]
        if mmr == "Placement":
            await(await ctx.send("Player %s needs to be given Placement MMR before they can receive a penalty!"
                           % (goodName))).delete(delay=15)
            return
        while int(mmr) - amount < 0:
            amount -= 1
        pens += amount
        await pHistory.update_cell(pHrow, 4, pens)
        e = discord.Embed(title="Penalty added")
        e.add_field(name="Player", value=goodName, inline=False)
        e.add_field(name="Penalty", value="-%d MMR:\n%d -> %d" % (amount, int(mmr), int(mmr)-amount), inline=False)
        if reason != "":
            e.add_field(name="Reason", value=reason, inline=False)
        await channel.send(embed=e)
        if channel == ctx.channel:
            try:
                await ctx.message.delete()
            except:
                pass
        else:
            await ctx.send("-%d penalty given to %s in %s"
                           % (amount, goodName, channel.mention))


    @update.command(aliases=['pl'])
    @commands.max_concurrency(number=1,wait=True)
    @commands.has_any_role("Administrator", "Updater", "Staff-S")
    async def place(self, ctx, rank, *, playername):
        if ctx.guild.id != self.config["server"]:
            await ctx.send("You cannot use this command in this server!")
            return
        #ranks = {"gold": 5500, "silver": 4000, "bronze": 2500, "iron": 1000}
        if rank.lower() not in place_MMRs.keys():
            await ctx.send("Please enter one of the following ranks: %s"
                           % (", ".join(ranks)))
            return
        placemmr = place_MMRs[rank.lower()]
        agc = await agcm.authorize()
        sh = await agc.open_by_key(SH_KEY)
        botSheet = await sh.worksheet("Bot")
        pHistory = await sh.worksheet("Player History")
        await botSheet.update_cell(pen_row, pen_cols[0], playername)
        info = await botSheet.batch_get(["%s:%s" % (get_strike_info[0], get_strike_info[1])])
        pHrow = info[0][0][2]
        mmr = info[0][0][7]
        if pHrow == "#N/A" or len(info[0][0]) < 8:
            await(await ctx.send("This player is not on the sheet; check your input and try again")).delete(delay=15)
            return
        goodName = info[0][0][8]
        
        if mmr != "Placement":
            await(await ctx.send("This player does not have Placement MMR! Check the sheet and try again")).delete(delay=15)
            return
        await pHistory.update_cell(pHrow, 7, placemmr)
        await ctx.message.delete()
        await ctx.send("Successfully placed %s in %s with %d MMR; make sure to give them the %s role in server!"
                       % (goodName, rank.lower(), placemmr, rank.lower()))

    async def processInstructions(self, ctx, tier, instructions):
        returnInstructions = {}
        if len(instructions) > 0:
            strings = instructions.split(",")
            for string in strings:
                idAmount = string.split()
                if len(idAmount) != 2:
                    await(await ctx.send("Your instruction `%s` needs to have 2 arguments separated by spaces"
                                   % string)).delete(delay=15)
                    raise Exception()
                try:
                    playerID = int(idAmount[0].strip())
                except:
                    await(await ctx.send("Your first argument in instruction `%s` is not an integer!" %
                                   string)).delete(delay=15)
                    raise Exception()
                if playerID < 1 or playerID > 12:
                    await(await ctx.send("Your first argument in instruction `%s` needs to be between 1 and 12"
                                   % string)).delete(delay=15)
                    raise Exception()
                try:
                    multiplier = float(idAmount[1].strip())
                except:
                    await(await ctx.send("Your second argument in instruction `%s` is not a float between 0-2!"
                                   % string)).delete(delay=15)
                    raise Exception()
                if multiplier < 0 or multiplier > 2:
                    await(await ctx.send("Your second argument in instruction `%s` is not a float between 0-2!"
                                   % string)).delete(delay=15)
                    raise Exception()
                returnInstructions[playerID] = multiplier
        return returnInstructions
        
        
def setup(bot):
    bot.add_cog(Updating(bot))
